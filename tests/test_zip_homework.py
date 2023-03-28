"""
Написать автотест, запаковывающий в zip-архив несколько разных файлов – pdf, xlsx, csv, – и кладущий его в папку
resources.
Реализовать чтение и проверку в ассертах содержимого каждого файла из архива, не распаковывая его.
"""
import csv
import io
import os
from zipfile import ZipFile

import pytest
from PyPDF2 import PdfReader
from openpyxl import load_workbook

file_names = ['username.csv',
              'docs-pytest-org-en-latest.pdf',
              'file_example_XLSX_10.xlsx']

resources_path = os.path.join(os.getcwd(), "..", "resources")
zip_path = f'{os.path.abspath(resources_path)}/output.zip'


@pytest.fixture
def delete_zip():
    if os.path.exists(zip_path):
        os.remove(zip_path)


@pytest.fixture
def zip_files(delete_zip):
    with ZipFile(zip_path, 'w') as zip_file:
        for file in file_names:
            file_path = os.path.join(os.getcwd(), "..", file)
            zip_file.write(file_path, file)


def test_zip(zip_files):
    compare_files_sizes()
    compare_xlsxs_rows_count()
    compare_pdfs_pages_count()
    compare_csvs_content()


def compare_files_sizes():
    for file in file_names:
        file_size = os.path.getsize(os.path.join(os.getcwd(), "..", file))

        with ZipFile(zip_path, 'r') as zip_file:
            file_size_in_zip = zip_file.getinfo(file).file_size

        assert file_size == file_size_in_zip


def compare_xlsxs_rows_count():
    xlsx_path = os.path.abspath(os.path.join(os.getcwd(), "..", file_names[2]))
    wb = load_workbook(xlsx_path)

    with ZipFile(zip_path, 'r') as zip_file:
        with zip_file.open(file_names[2], 'r') as excel_file1:
            excel_bytes = io.BytesIO(excel_file1.read())
            wb_zip = load_workbook(excel_bytes)

        for sheet, sheet_zip in zip(wb.worksheets, wb_zip.worksheets):
            assert sheet.max_row == sheet_zip.max_row

    wb.close()
    wb_zip.close()


def compare_pdfs_pages_count():
    pdf_path = os.path.abspath(os.path.join(os.getcwd(), "..", file_names[1]))

    reader = PdfReader(pdf_path)
    numbers_of_pages = len(reader.pages)

    with ZipFile(zip_path, 'r') as zip_file:
        with zip_file.open(file_names[1], 'r') as pdf_zip_file:
            pdf_zip_reader = PdfReader(pdf_zip_file)
            numbers_of_pages_in_zip = len(pdf_zip_reader.pages)

    assert numbers_of_pages == numbers_of_pages_in_zip


def compare_csvs_content():
    csv_path = os.path.abspath(os.path.join(os.getcwd(), "..", file_names[0]))

    with open(csv_path, 'r') as csv_file:
        csv_reader = csv.reader(csv_file)

        with ZipFile(zip_path, 'r') as zip_file:
            file_contents = zip_file.read(file_names[0])
            file_contents_string = file_contents.decode('utf-8')
            csv_string_reader = csv.reader(file_contents_string.splitlines())

            for row1, row2 in zip(csv_reader, csv_string_reader):
                assert row1 == row2
